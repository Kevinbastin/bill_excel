#!/usr/bin/env python3
"""
INTEGRATION: PPStructure → InvoiceTableExtractor → Excel
"""

import sys
import cv2
from typing import List, Dict, Any
from paddleocr import PPStructure
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .invoice_extractor import InvoiceTableExtractor


class InvoiceOCRPipeline:
    def __init__(self):
        self.engine = PPStructure(use_cuda=False, lang="ch", recovery=False, layout=True, table=True, ocr=True)
        self.extractor = InvoiceTableExtractor(y_threshold=20)

    def run(self, image_path: str, output_xlsx: str) -> bool:
        print(f"\n{'='*80}\nINVOICE OCR → EXTRACTION → EXCEL\n{'='*80}\n")
        
        print(f"[1/5] Reading image: {image_path}")
        img = cv2.imread(image_path)
        if img is None:
            print(f"❌ Failed to read image")
            return False
        print(f"✓ Image loaded: {img.shape[1]}×{img.shape[0]}")

        print(f"\n[2/5] Running PP-Structure OCR...")
        result = self.engine(img, return_ocr_result_in_table=True)
        print(f"✓ OCR complete")

        print(f"\n[3/5] Extracting table cells...")
        cells = self._extract_cells_from_result(result)
        if not cells:
            print(f"❌ No cells found")
            return False
        print(f"✓ Found {len(cells)} cells")

        print(f"\n[4/5] Reconstructing table structure...")
        grid = self.extractor.extract(cells)
        if not grid:
            print(f"❌ Failed to extract grid")
            return False
        print(f"✓ Extracted {len(grid)} rows, {len(grid[0]) if grid else 0} columns")

        print(f"\nPreview of extracted data:")
        for i, row in enumerate(grid[:6]):
            print(f"  Row {i}: {row}")

        print(f"\n[5/5] Exporting to Excel: {output_xlsx}")
        success = self._export_to_excel(grid, output_xlsx)
        if success:
            print(f"✓ Excel file created successfully")
            return True
        return False

    @staticmethod
    def _extract_cells_from_result(result: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        cells = []
        for region in result:
            if region.get("type") != "table":
                continue
            res = region.get("res", {}) or {}
            rec_res = res.get("rec_res", []) or []
            boxes = res.get("boxes", None)
            cell_bbox = res.get("cell_bbox", None)
            
            if isinstance(boxes, (list, tuple)) and len(boxes) == len(rec_res):
                for item, b in zip(rec_res, boxes):
                    txt, score = InvoiceOCRPipeline._extract_text_score(item)
                    if not txt:
                        continue
                    xyxy = InvoiceOCRPipeline._box_to_xyxy(b)
                    if not xyxy:
                        continue
                    x1, y1, x2, y2 = xyxy
                    cells.append({"text": txt, "score": float(score or 0), "x1": x1, "y1": y1, "x2": x2, "y2": y2})
            
            elif isinstance(cell_bbox, (list, tuple)) and len(cell_bbox) == len(rec_res):
                for item, b in zip(rec_res, cell_bbox):
                    txt, score = InvoiceOCRPipeline._extract_text_score(item)
                    if not txt:
                        continue
                    xyxy = InvoiceOCRPipeline._box_to_xyxy(b)
                    if not xyxy:
                        continue
                    x1, y1, x2, y2 = xyxy
                    cells.append({"text": txt, "score": float(score or 0), "x1": x1, "y1": y1, "x2": x2, "y2": y2})
        return cells

    @staticmethod
    def _extract_text_score(item: Any) -> tuple:
        txt, score = "", None
        if isinstance(item, (list, tuple)) and len(item) >= 1:
            txt = str(item[0]).strip() if item[0] else ""
            score = float(item[1]) if len(item) >= 2 and item[1] else None
        elif isinstance(item, dict):
            txt = str(item.get("text") or item.get("transcription") or "").strip()
            score = item.get("score") or item.get("confidence")
            try:
                score = float(score) if score else None
            except:
                score = None
        return txt, score

    @staticmethod
    def _box_to_xyxy(b: Any) -> Any:
        try:
            if not isinstance(b, (list, tuple)):
                return None
            b = list(b)
            if len(b) == 4 and all(isinstance(v, (int, float)) for v in b):
                x1, y1, x2, y2 = b
                return int(min(x1, x2)), int(min(y1, y2)), int(max(x1, x2)), int(max(y1, y2))
            if len(b) == 8 and all(isinstance(v, (int, float)) for v in b):
                xs = b[0::2]
                ys = b[1::2]
                return int(min(xs)), int(min(ys)), int(max(xs)), int(max(ys))
            if len(b) == 4 and isinstance(b[0], (list, tuple)) and len(b[0]) == 2:
                xs = [p[0] for p in b]
                ys = [p[1] for p in b]
                return int(min(xs)), int(min(ys)), int(max(xs)), int(max(ys))
            return None
        except:
            return None

    @staticmethod
    def _export_to_excel(grid: List[List[str]], output_path: str) -> bool:
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Invoice"
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for row_idx, row_data in enumerate(grid, start=1):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    if row_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            for col_idx in range(1, len(grid[0]) + 1 if grid else 1):
                max_length = 0
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                for row in grid:
                    if col_idx <= len(row):
                        cell_len = len(str(row[col_idx - 1]))
                        max_length = max(max_length, cell_len)
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width
            
            for row_idx in range(1, len(grid) + 1):
                ws.row_dimensions[row_idx].height = 30
            
            wb.save(output_path)
            return True
        except Exception as e:
            print(f"Excel export error: {e}")
            return False


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_export.py <image.jpg> [output.xlsx]")
        sys.exit(1)
    image_path = sys.argv[1]
    output_xlsx = sys.argv[2] if len(sys.argv) > 2 else "extracted_invoice.xlsx"
    pipeline = InvoiceOCRPipeline()
    success = pipeline.run(image_path, output_xlsx)
    sys.exit(0 if success else 1)
