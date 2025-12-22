from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional
import logging
import json

logger = logging.getLogger(__name__)


class ExcelWriter:
    @staticmethod
    def get_styles() -> Dict[str, Any]:
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)

        table_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        table_header_font = Font(color="FFFFFF", bold=True, size=11)

        data_font = Font(size=10)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        confidence_high = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        confidence_med = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        confidence_low = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")

        return {
            "header_fill": header_fill,
            "header_font": header_font,
            "table_header_fill": table_header_fill,
            "table_header_font": table_header_font,
            "data_font": data_font,
            "border": border,
            "confidence_high": confidence_high,
            "confidence_med": confidence_med,
            "confidence_low": confidence_low,
        }

    @staticmethod
    def safe_str(x: Any) -> str:
        if x is None:
            return ""
        if isinstance(x, (int, float)):
            return str(x)
        return str(x).strip() if hasattr(x, "strip") else str(x)

    @staticmethod
    def _normalize_grid(grid: List[List[Any]]) -> List[List[str]]:
        out: List[List[str]] = []
        max_cols = max((len(r) for r in grid), default=0)
        for r in grid or []:
            row = [ExcelWriter.safe_str(v) for v in (r or [])]
            if len(row) < max_cols:
                row += [""] * (max_cols - len(row))
            out.append(row)
        return out

    @staticmethod
    def _row_text(row: List[str]) -> str:
        return " ".join((c or "").strip().lower() for c in (row or []) if (c or "").strip())

    @staticmethod
    def _is_placeholder_header_row(row: List[str]) -> bool:
        txt = ExcelWriter._row_text(row)
        return "col1" in txt and "col2" in txt

    @staticmethod
    def _is_summary_or_totals_row(row: List[str]) -> bool:
        """
        Detect summary/footer rows like:
        - Subtotal
        - Tax Rate / Tax / GST / VAT
        - TOTAL / Grand Total / Amount Due / Balance Due
        """
        txt = ExcelWriter._row_text(row)
        stop_keys = (
            "subtotal", "sub total", "sub-total",
            "tax rate", "tax", "gst", "vat", "igst", "cgst", "sgst",
            "total", "grand total", "total due", "amount due", "balance due",
            "round off", "roundoff",
            "shipping", "handling", "delivery",
            "discount",
            "net amount", "gross amount",
        )
        return any(k in txt for k in stop_keys)

    @staticmethod
    def _trim_empty_cols(grid: List[List[str]]) -> List[List[str]]:
        if not grid:
            return grid
        max_cols = max((len(r) for r in grid), default=0)
        last_used = -1
        for c in range(max_cols):
            if any((c < len(r) and (r[c] or "").strip()) for r in grid):
                last_used = c
        if last_used < 0:
            return grid
        return [r[: last_used + 1] for r in grid]

    @staticmethod
    def _parse_html_table(html: str) -> Optional[List[List[str]]]:
        if not html or "<table" not in html.lower():
            return None
        try:
            import pandas as pd
            from io import StringIO
            dfs = pd.read_html(StringIO(html))
            if not dfs:
                return None
            df = dfs[0]
            header = [str(c).strip() for c in df.columns]
            rows = [header]
            for _, r in df.iterrows():
                rows.append([str(v).strip() for v in r.values])
            return rows if len(rows) > 1 else None
        except Exception as e:
            logger.warning(f"HTML table parse skipped: {type(e).__name__}: {e}")
            return None

    @staticmethod
    def _find_header_row(grid: List[List[str]], max_scan: int = 30) -> Optional[int]:
        """
        Enhanced header detection for invoices like inv123.jpg which has:
        SL | Item Description | Price | Qty | Total
        """
        keywords = (
            # ID/Serial columns
            "sl", "sl.", "s.no", "no", "no.", "id", "sr", "sr.",
            # Description
            "description", "desc", "item", "product", "particulars",
            # Quantity
            "qty", "quantity", "qnty", "units",
            # Unit of Measure
            "uom", "um", "unit",
            # Price
            "price", "rate", "net price", "unit price", "amount",
            # Total
            "total", "net worth", "gross", "gross worth",
            # Tax columns
            "vat", "gst", "hsn", "sac", "igst", "cgst", "sgst", "cess",
        )
        best_i = None
        best_hits = 0
        for i in range(min(max_scan, len(grid))):
            row = grid[i]
            if not row:
                continue
            if ExcelWriter._is_placeholder_header_row(row):
                continue
            txt = ExcelWriter._row_text(row)
            if not txt:
                continue
            hits = sum(1 for k in keywords if k in txt)
            if hits > best_hits:
                best_hits = hits
                best_i = i
        # Lower threshold for simple invoices with fewer columns
        return best_i if best_i is not None and best_hits >= 2 else None

    @staticmethod
    def _crop_to_table(grid: List[List[str]]) -> Optional[List[List[str]]]:
        """
        Crop grid to:
        - Start: header row
        - End: BEFORE first summary/total row (Subtotal, Tax Rate, TOTAL etc)
        """
        if not grid or len(grid) < 2:
            return None
        h = ExcelWriter._find_header_row(grid)
        if h is None:
            return None

        cropped: List[List[str]] = [grid[h]]  # Start with header
        
        for i in range(h + 1, len(grid)):
            row = grid[i]
            # Stop BEFORE summary rows
            if ExcelWriter._is_summary_or_totals_row(row):
                logger.debug(f"Stopping at summary row {i}: {ExcelWriter._row_text(row)[:50]}")
                break
            cropped.append(row)

        cropped = ExcelWriter._trim_empty_cols(cropped)
        return cropped if len(cropped) >= 2 else None

    @staticmethod
    def autofit_columns(ws, min_width: int = 12, max_width: int = 60) -> None:
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is not None and str(v).strip():
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))

    @staticmethod
    def _write_grid_block(ws, start_row: int, title: str, grid: List[List[str]], styles: Dict) -> int:
        grid = ExcelWriter._normalize_grid(grid)
        ncols = max((len(r) for r in grid), default=1)
        end_col = get_column_letter(max(1, ncols))

        # Title row
        ws.merge_cells(f"A{start_row}:{end_col}{start_row}")
        title_cell = ws[f"A{start_row}"]
        title_cell.value = title
        title_cell.font = styles["table_header_font"]
        title_cell.fill = styles["table_header_fill"]
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        start_row += 1

        # Header row
        for c, v in enumerate(grid[0], 1):
            cell = ws.cell(row=start_row, column=c)
            cell.value = ExcelWriter.safe_str(v)
            cell.font = styles["table_header_font"]
            cell.fill = styles["table_header_fill"]
            cell.border = styles["border"]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        start_row += 1

        # Data rows
        for r in grid[1:]:
            for c, v in enumerate(r, 1):
                cell = ws.cell(row=start_row, column=c)
                cell.value = ExcelWriter.safe_str(v)
                cell.font = styles["data_font"]
                cell.border = styles["border"]
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            start_row += 1

        return start_row + 1

    @staticmethod
    def write_invoice_details_sheet(ws, extracted_data: Dict, styles: Dict) -> None:
        row = 1
        ws.merge_cells(f"A{row}:F{row}")
        title_cell = ws[f"A{row}"]
        title_cell.value = "INVOICE DETAILS"
        title_cell.font = styles["header_font"]
        title_cell.fill = styles["header_fill"]
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 2

        fields = extracted_data.get("fields") or {}
        confidence = float(extracted_data.get("confidence_score") or 0.0)

        field_pairs = [
            ("Invoice Number", fields.get("invoice_number") or "N/A"),
            ("Date", fields.get("date") or "N/A"),
            ("Vendor", fields.get("vendor") or "N/A"),
            ("Total Amount", fields.get("total_amount") or "N/A"),
            ("GSTIN", fields.get("gstin") or "N/A"),
            ("PO Number", fields.get("po_number") or "N/A"),
            ("Due Date", fields.get("due_date") or "N/A"),
        ]
        for label, value in field_pairs:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True, size=11)
            ws[f"B{row}"] = ExcelWriter.safe_str(value)
            ws[f"B{row}"].font = styles["data_font"]
            row += 1

        ws[f"A{row}"] = "OCR Confidence Score"
        ws[f"A{row}"].font = Font(bold=True, size=11)
        conf_cell = ws[f"B{row}"]
        conf_cell.value = f"{confidence:.1f}%"
        conf_cell.font = Font(bold=True, size=11)
        if confidence >= 90:
            conf_cell.fill = styles["confidence_high"]
        elif confidence >= 75:
            conf_cell.fill = styles["confidence_med"]
        else:
            conf_cell.fill = styles["confidence_low"]

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40

    @staticmethod
    def write_line_items_sheet(ws, extracted_data: Dict, styles: Dict) -> None:
        row = 1
        tables = extracted_data.get("tables") or []
        if not tables:
            ws["A1"] = "No tables detected in invoice"
            ws["A1"].font = Font(italic=True, color="FF0000")
            return

        for table_idx, table in enumerate(tables, 1):
            # Priority 1: canonical_rows / canonicalrows
            grid = table.get("canonical_rows") or table.get("canonicalrows") or table.get("raw_grid")

            # Priority 2: HTML parsing (if available)
            if grid is None and table.get("html"):
                grid = ExcelWriter._parse_html_table(table.get("html"))

            # Priority 3: Fallback to text tokens
            if grid is None and table.get("text"):
                text_items = table.get("text", [])
                if text_items:
                    grid = [["Item", "Text"]] + [[str(i), str(t)] for i, t in enumerate(text_items, 1)]

            if not grid:
                ws[f"A{row}"] = f"TABLE {table_idx}: No data extracted"
                ws[f"A{row}"].font = Font(italic=True, color="FF9900")
                row += 2
                continue

            # Normalize and crop
            grid = ExcelWriter._normalize_grid(grid)
            cropped = ExcelWriter._crop_to_table(grid)

            if cropped:
                row = ExcelWriter._write_grid_block(
                    ws, row, f"TABLE {table_idx} - LINE ITEMS", cropped, styles
                )
            else:
                # Fallback: write raw grid if cropping fails
                row = ExcelWriter._write_grid_block(
                    ws, row, f"TABLE {table_idx} - LINE ITEMS (RAW)", grid, styles
                )

        ExcelWriter.autofit_columns(ws)

    @staticmethod
    def write_json_sheet(ws, extracted_data: Dict, styles: Dict) -> None:
        row = 1
        ws.merge_cells(f"A{row}:B{row+20}")
        json_cell = ws[f"A{row}"]
        json_str = json.dumps(extracted_data, indent=2, default=str)[:5000]
        json_cell.value = json_str
        json_cell.alignment = Alignment(wrap_text=True, vertical="top")
        json_cell.font = Font(name="Courier New", size=8)
        ws.column_dimensions["A"].width = 60
        ws.column_dimensions["B"].width = 30

    @staticmethod
    def write_invoice_to_excel(extracted_data: Dict, output_path: str) -> str:
        wb = Workbook()
        styles = ExcelWriter.get_styles()

        ws1 = wb.active
        ws1.title = "Invoice Details"
        ExcelWriter.write_invoice_details_sheet(ws1, extracted_data, styles)

        ws2 = wb.create_sheet("Line Items")
        ExcelWriter.write_line_items_sheet(ws2, extracted_data, styles)

        ws3 = wb.create_sheet("Raw JSON")
        ExcelWriter.write_json_sheet(ws3, extracted_data, styles)

        wb.save(output_path)
        logger.info(f"✅ Excel file saved: {output_path}")
        return output_path
